import openpyxl


class ExcelMethods_class:

    def getRowCount(self, file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_row


    def  read_data(self, file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value


    def write_data(self, file, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)